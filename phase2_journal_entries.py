"""
Phase 2: 회계 전표 변환 + 분개 자동 생성
- 성진이 만든 je_rules 엑셀(결제수단별 매핑, 카테고리별 원가율)을 읽어서
- 원본 거래 데이터에 통제 필드(전표번호, 승인자, Posting Date 등)를 붙이고
- 매출 분개 + 매출원가 분개를 자동 생성한 뒤
- 감사/수작업 검토용 엑셀로 export 한다.
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def load_payment_rules(path: str) -> dict:
    """'결제수단별 매출 분개매핑' 시트를 읽어 {Payment Method: {debit, credit}} 딕셔너리로 반환."""
    df = pd.read_excel(path, sheet_name="결제수단별 매출 분개매핑")
    return {
        row["Payment Method (원본 컬럼값)"]: {
            "debit": row["차변 계정"],
            "credit": row["대변 계정(기본)"],
        }
        for _, row in df.iterrows()
    }


def load_category_rules(path: str) -> dict:
    """'카테고리별 원가율 및 세부계정' 시트를 읽어 {Category: {...}} 딕셔너리로 반환."""
    df = pd.read_excel(path, sheet_name="카테고리별 원가율 및 세부계정")
    rules = {}
    for _, row in df.iterrows():
        cogs_rate = float(str(row["추정 원가율"]).replace("%", "")) / 100
        rules[row["Category (원본 컬럼값)"]] = {
            "revenue_account": row["대변(수익) 세부계정명"],
            "cogs_rate": cogs_rate,
            "cogs_debit": row["차변(비용) 계정"],
            "cogs_credit": row["대변(자산감소) 계정"],
        }
    return rules


def transform_to_voucher(df: pd.DataFrame, seed: int = 42) -> pd.DataFrame:
    """
    원본 거래 데이터에 회계 통제 필드를 추가해서 '전표' 형태로 만든다.
    추가 필드: Voucher No, Prepared By, Approved By, Posting Date, Void Flag

    승인자는 "김성진" 한 명으로 고정하고, 기안자는 "김경석"/"홍성현"/"장준호" 중
    무작위로 배정한다. 승인자 풀과 기안자 풀을 완전히 분리해서, 이 단계에서는
    직무분리(기안자=승인자) 위반이 우연히도 발생하지 않는 '깨끗한 기준선'으로 만든다.
    직무분리 위반 같은 통제 미비 시나리오는 Phase 5에서 의도적으로 삽입한다.
    """
    rng = np.random.default_rng(seed)
    voucher = df.copy()

    voucher["Voucher No"] = ["V" + str(i).zfill(6) for i in range(1, len(voucher) + 1)]

    preparer_pool = ["김경석", "홍성현", "장준호"]
    voucher["Prepared By"] = rng.choice(preparer_pool, size=len(voucher))
    voucher["Approved By"] = "김성진"

    voucher["Transaction Date"] = pd.to_datetime(voucher["Transaction Date"])
    delay_days = rng.integers(0, 4, size=len(voucher))
    voucher["Posting Date"] = voucher["Transaction Date"] + pd.to_timedelta(delay_days, unit="D")

    voucher["Void Flag"] = False
    return voucher


def generate_journal_entries(voucher: pd.DataFrame, payment_rules: dict, category_rules: dict) -> pd.DataFrame:
    """전표 한 건당 분개 2쌍(매출 + 매출원가)을 생성해 long format DataFrame으로 반환."""
    lines = []
    skipped = 0

    for _, row in voucher.iterrows():
        pm, cat, amount = row["Payment Method"], row["Category"], row["Total Spent"]

        if pd.isna(amount) or pd.isna(pm) or pd.isna(cat) or pm not in payment_rules or cat not in category_rules:
            skipped += 1
            continue

        pay = payment_rules[pm]
        cr = category_rules[cat]
        cogs = round(amount * cr["cogs_rate"], 2)

        common = {"Voucher No": row["Voucher No"], "Prepared By": row["Prepared By"],
                  "Approved By": row["Approved By"], "Posting Date": row["Posting Date"]}

        lines += [
            {**common, "JE Type": "매출", "Account": pay["debit"], "Debit": amount, "Credit": 0},
            {**common, "JE Type": "매출", "Account": cr["revenue_account"], "Debit": 0, "Credit": amount},
            {**common, "JE Type": "매출원가", "Account": cr["cogs_debit"], "Debit": cogs, "Credit": 0},
            {**common, "JE Type": "매출원가", "Account": cr["cogs_credit"], "Debit": 0, "Credit": cogs},
        ]

    print(f"분개 생성 건너뜀 (결측치 등): {skipped}건")
    return pd.DataFrame(lines)


def export_je_to_excel(je: pd.DataFrame, output_path: str) -> None:
    """분개 리스트를 감사/수작업 검토용 엑셀로 export (검토자/검토일/비고 열 추가, 서식 적용)."""
    export_df = je.copy()
    export_df["검토자"] = ""
    export_df["검토일"] = ""
    export_df["비고"] = ""
    export_df.to_excel(output_path, index=False, sheet_name="Journal Entries")

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Journal Entries"]

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Posting Date 컬럼 위치를 헤더에서 찾아서 날짜 서식 지정 (### 오류 방지)
    headers = [c.value for c in ws[1]]
    date_col_idx = headers.index("Posting Date") + 1 if "Posting Date" in headers else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if date_col_idx and cell.column == date_col_idx:
                cell.number_format = "YYYY-MM-DD"

    widths = [12, 12, 12, 14, 10, 22, 10, 10, 12, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


if __name__ == "__main__":
    import sys

    DATA_PATH = sys.argv[1] if len(sys.argv) > 1 else "data/retail_sales.csv"
    RULES_PATH = sys.argv[2] if len(sys.argv) > 2 else "data/Je_rules_매핑_통합본.xlsx"
    OUTPUT_PATH = "journal_entries.xlsx"

    df = pd.read_csv(DATA_PATH)
    payment_rules = load_payment_rules(RULES_PATH)
    category_rules = load_category_rules(RULES_PATH)

    voucher = transform_to_voucher(df)
    je = generate_journal_entries(voucher, payment_rules, category_rules)
    export_je_to_excel(je, OUTPUT_PATH)

    print(f"\n생성된 분개 라인 수: {len(je)}")
    print(f"차변 합계: {je['Debit'].sum():,.0f} / 대변 합계: {je['Credit'].sum():,.0f}")
    print(f"엑셀 저장 완료: {OUTPUT_PATH}")
