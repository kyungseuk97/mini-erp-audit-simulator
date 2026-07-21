"""
Phase 4: GL -> TB -> F/S 자동 집계 (2024년 데이터 기준, 완전한 재무상태표 포함)

- 2024년 거래만 필터링 (Transaction Date 기준)
- 가상의 기초 재무상태표(2024-01-01)를 가정하여 기말 재무상태표를 완성한다.
  (원본 데이터는 매출 사이클만 있어 부채/자본 변동 거래가 없으므로,
   매입채무/자본금은 기초 그대로 유지되고 이익잉여금만 당기순이익만큼 증가한다.)
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 가상 기초 재무상태표 (2024-01-01 기준, 가정치)
OPENING_BALANCE = {
    "현금": 5_000_000,
    "매출채권(카드사)": 2_000_000,
    "미결제예치금": 1_000_000,
    "상품(재고자산)": 8_000_000,
    "매입채무": 3_000_000,
    "자본금": 10_000_000,
    "이익잉여금": 3_000_000,
}


def filter_year(je: pd.DataFrame, year: int, date_col: str = "Transaction Date") -> pd.DataFrame:
    """분개 데이터를 특정 연도(거래일자 기준)로 필터링한다."""
    return je[je[date_col].dt.year == year].copy()


def generate_gl(je: pd.DataFrame) -> pd.DataFrame:
    """분개를 계정과목별로 집계해서 총계정원장(GL)을 만든다."""
    gl = je.groupby("Account", as_index=False)[["Debit", "Credit"]].sum()
    gl["Balance"] = gl["Debit"] - gl["Credit"]
    return gl


def generate_tb(gl: pd.DataFrame) -> tuple[pd.DataFrame, bool]:
    """GL을 시산표(TB) 형태로 변환하고 차대변 일치 여부를 검증한다."""
    tb = gl.copy()
    tb["TB_Debit"] = tb["Balance"].apply(lambda x: x if x > 0 else 0)
    tb["TB_Credit"] = tb["Balance"].apply(lambda x: -x if x < 0 else 0)
    balanced = abs(tb["TB_Debit"].sum() - tb["TB_Credit"].sum()) < 0.01
    return tb[["Account", "TB_Debit", "TB_Credit"]], balanced


def generate_income_statement(gl: pd.DataFrame) -> tuple[pd.DataFrame, float]:
    """GL에서 매출/매출원가 계정을 뽑아 손익계산서를 만들고, 당기순이익을 반환한다."""
    revenue = gl[gl["Account"].str.startswith("매출-")]["Credit"].sum()
    cogs = gl[gl["Account"].str.startswith("매출원가-")]["Debit"].sum()
    net_income = revenue - cogs
    income_statement = pd.DataFrame([
        {"항목": "매출액", "금액": revenue},
        {"항목": "매출원가", "금액": -cogs},
        {"항목": "매출총이익(당기순이익)", "금액": net_income},
    ])
    return income_statement, net_income


def generate_balance_sheet(gl: pd.DataFrame, opening: dict, net_income: float) -> pd.DataFrame:
    """가상 기초 재무상태표 + 당기 GL 증감을 반영해 기말 재무상태표를 만든다."""
    def bal(names):
        if isinstance(names, str):
            names = [names]
        return gl.loc[gl["Account"].isin(names), "Balance"].sum()

    inventory_accounts = [a for a in gl["Account"] if a.startswith("상품-")]

    closing = {
        "현금": opening["현금"] + bal("현금"),
        "매출채권(카드사)": opening["매출채권(카드사)"] + bal("매출채권(카드사)"),
        "미결제예치금": opening["미결제예치금"] + bal("미결제예치금"),
        "상품(재고자산)": opening["상품(재고자산)"] + bal(inventory_accounts),
        "매입채무": opening["매입채무"],
        "자본금": opening["자본금"],
        "이익잉여금": opening["이익잉여금"] + net_income,
    }

    rows = []
    rows.append({"구분": "[자산]", "계정과목": "", "금액": None})
    for k in ["현금", "매출채권(카드사)", "미결제예치금", "상품(재고자산)"]:
        rows.append({"구분": "자산", "계정과목": k, "금액": closing[k]})
    assets_total = closing["현금"] + closing["매출채권(카드사)"] + closing["미결제예치금"] + closing["상품(재고자산)"]
    rows.append({"구분": "자산총계", "계정과목": "", "금액": assets_total})

    rows.append({"구분": "[부채]", "계정과목": "", "금액": None})
    rows.append({"구분": "부채", "계정과목": "매입채무", "금액": closing["매입채무"]})

    rows.append({"구분": "[자본]", "계정과목": "", "금액": None})
    rows.append({"구분": "자본", "계정과목": "자본금", "금액": closing["자본금"]})
    rows.append({"구분": "자본", "계정과목": "이익잉여금", "금액": closing["이익잉여금"]})
    liab_equity_total = closing["매입채무"] + closing["자본금"] + closing["이익잉여금"]
    rows.append({"구분": "부채와자본총계", "계정과목": "", "금액": liab_equity_total})

    rows.append({"구분": "일치 여부", "계정과목": "", "금액": "일치함" if abs(assets_total - liab_equity_total) < 0.01 else "불일치"})

    return pd.DataFrame(rows)


def style_sheet(ws, header_row=1):
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.font = Font(name="Arial", size=10)
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 4)
    ws.freeze_panes = f"A{header_row + 1}"


def export_fs_to_excel(gl, tb, income_statement, balance_sheet, balanced, output_path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        gl.to_excel(writer, sheet_name="GL", index=False)
        tb.to_excel(writer, sheet_name="TB", index=False)
        income_statement.to_excel(writer, sheet_name="손익계산서", index=False)
        balance_sheet.to_excel(writer, sheet_name="재무상태표", index=False)

    wb = openpyxl.load_workbook(output_path)
    for name in wb.sheetnames:
        style_sheet(wb[name])

    ws_tb = wb["TB"]
    note = ws_tb.cell(row=ws_tb.max_row + 2, column=1, value=f"차대변 일치 여부: {'일치함' if balanced else '불일치! 확인 필요'}")
    note.font = Font(bold=True, color="008000" if balanced else "FF0000")

    wb.save(output_path)


if __name__ == "__main__":
    import sys
    from phase2_journal_entries import load_payment_rules, load_category_rules, transform_to_voucher

    DATA_PATH = sys.argv[1] if len(sys.argv) > 1 else "data/retail_sales.csv"
    RULES_PATH = sys.argv[2] if len(sys.argv) > 2 else "data/Je_rules_매핑_통합본.xlsx"
    YEAR = int(sys.argv[3]) if len(sys.argv) > 3 else 2024

    df = pd.read_csv(DATA_PATH)
    payment_rules = load_payment_rules(RULES_PATH)
    category_rules = load_category_rules(RULES_PATH)
    voucher = transform_to_voucher(df)

    # generate_journal_entries에 Transaction Date도 같이 담아야 하므로 phase2 함수를 살짝 확장해서 재사용
    def generate_journal_entries_with_date(voucher, pay_rules, cat_rules):
        lines = []
        for _, row in voucher.iterrows():
            pm, cat, amount = row["Payment Method"], row["Category"], row["Total Spent"]
            if pd.isna(amount) or pd.isna(pm) or pd.isna(cat) or pm not in pay_rules or cat not in cat_rules:
                continue
            pay, cr = pay_rules[pm], cat_rules[cat]
            cogs = round(amount * cr["cogs_rate"], 2)
            common = {"Voucher No": row["Voucher No"], "Transaction Date": row["Transaction Date"], "Posting Date": row["Posting Date"]}
            lines += [
                {**common, "JE Type": "매출", "Account": pay["debit"], "Debit": amount, "Credit": 0},
                {**common, "JE Type": "매출", "Account": cr["revenue_account"], "Debit": 0, "Credit": amount},
                {**common, "JE Type": "매출원가", "Account": cr["cogs_debit"], "Debit": cogs, "Credit": 0},
                {**common, "JE Type": "매출원가", "Account": cr["cogs_credit"], "Debit": 0, "Credit": cogs},
            ]
        return pd.DataFrame(lines)

    je_all = generate_journal_entries_with_date(voucher, payment_rules, category_rules)
    je_year = filter_year(je_all, YEAR)

    gl = generate_gl(je_year)
    tb, balanced = generate_tb(gl)
    income_statement, net_income = generate_income_statement(gl)
    balance_sheet = generate_balance_sheet(gl, OPENING_BALANCE, net_income)

    export_fs_to_excel(gl, tb, income_statement, balance_sheet, balanced, "gl_tb_fs.xlsx")
    print(f"{YEAR}년 기준 GL/TB/F&S 생성 완료. 차대변 일치: {balanced}")
    print(f"당기순이익: {net_income:,.0f}")
